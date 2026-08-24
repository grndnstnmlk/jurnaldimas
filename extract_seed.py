import openpyxl
import json
import sqlite3
import os
from datetime import datetime

wb_data = openpyxl.load_workbook('JURNAL KEUANGAN (fix).xlsm', data_only=True)

# 1. Customer map from Nota Transaksi
ws_nota = wb_data['Nota Transaksi']
customer_map = {}
for r in range(9, 50):
    full_name = ws_nota.cell(r, 10).value
    code = ws_nota.cell(r, 11).value
    if full_name or code:
        c_code = str(code).strip() if code else ''
        c_name = str(full_name).strip() if full_name else c_code
        if c_code:
            customer_map[c_code] = c_name

# Customers in Master Data
ws_md = wb_data['Master Data']
all_customer_codes = []
for c in range(3, ws_md.max_column + 1):
    val = ws_md.cell(2, c).value
    if val and str(val).strip() not in ['Total Produk', '120']:
        c_code = str(val).strip()
        if c_code not in all_customer_codes:
            all_customer_codes.append(c_code)

customers_list = []
for code in all_customer_codes:
    name = customer_map.get(code, code)
    customers_list.append({
        'code': code,
        'name': name,
        'phone': '',
        'address': ''
    })

# 2. Products & Modal from Data Produk
ws_dp = wb_data['Data Produk']
product_cost_map = {}
for r in range(9, ws_dp.max_row + 1):
    prod = ws_dp.cell(r, 1).value
    cost = ws_dp.cell(r, 2).value
    if prod:
        p_name = str(prod).strip()
        # Modal is in thousands, e.g. 58 = 58,000
        try:
            c_val = float(cost) * 1000 if cost is not None else 0
        except:
            c_val = 0
        product_cost_map[p_name] = int(c_val)

# 3. Products list & Customer Prices from Master Data
products_list = []
pricing_matrix = [] # list of {product, customer_code, sell_price}

for r in range(3, ws_md.max_row + 1):
    prod = ws_md.cell(r, 2).value
    if prod:
        p_name = str(prod).strip()
        modal = product_cost_map.get(p_name, 0)
        
        products_list.append({
            'name': p_name,
            'category': 'Rokok',
            'modal_price': modal,
            'default_price': modal # fallback
        })
        
        for idx, cust_code in enumerate(all_customer_codes):
            c_col = idx + 3
            val = ws_md.cell(r, c_col).value
            try:
                price = int(float(val) * 1000) if val is not None and float(val) > 0 else 0
            except:
                price = 0
            
            pricing_matrix.append({
                'product_name': p_name,
                'customer_code': cust_code,
                'sell_price': price
            })

# 4. Stock Opname
ws_so = wb_data['Stock Opname']
stock_list = {}
for r in range(7, ws_so.max_row + 1):
    prod = ws_so.cell(r, 2).value
    awal = ws_so.cell(r, 3).value or 0
    in_qty = ws_so.cell(r, 4).value or 0
    out_qty = ws_so.cell(r, 5).value or 0
    akhir = ws_so.cell(r, 6).value or 0
    if prod:
        p_name = str(prod).strip()
        try:
            stock_list[p_name] = {
                'stok_awal': int(awal),
                'stok_in': int(in_qty),
                'stok_out': int(out_qty),
                'stok_akhir': int(akhir)
            }
        except:
            stock_list[p_name] = {'stok_awal': 0, 'stok_in': 0, 'stok_out': 0, 'stok_akhir': 0}

# 5. Transactions from Nota Transaksi
transactions_raw = []
for r in range(11, ws_nota.max_row + 1):
    tgl = ws_nota.cell(r, 2).value
    buyer = ws_nota.cell(r, 3).value
    prod = ws_nota.cell(r, 4).value
    qty = ws_nota.cell(r, 5).value
    subtotal = ws_nota.cell(r, 6).value
    
    if any([tgl, buyer, prod, qty, subtotal]):
        date_str = ''
        if isinstance(tgl, datetime):
            date_str = tgl.strftime('%Y-%m-%d')
        elif tgl:
            date_str = str(tgl).split(' ')[0]
        else:
            date_str = '2026-08-06'
            
        b_code = str(buyer).strip() if buyer else 'UMUM'
        p_name = str(prod).strip() if prod else ''
        q_val = int(qty) if qty else 1
        sub_val = int(float(subtotal)) if subtotal and str(subtotal) != 'Belum ada harga' else 0
        
        # Calculate unit price
        unit_price = sub_val // q_val if q_val > 0 else sub_val
        
        transactions_raw.append({
            'date': date_str,
            'customer_code': b_code,
            'product_name': p_name,
            'qty': q_val,
            'unit_price': unit_price,
            'subtotal': sub_val
        })

# Group transactions into invoices (Notas) by date and customer
invoices_grouped = {}
for t in transactions_raw:
    key = f"{t['date']}_{t['customer_code']}"
    if key not in invoices_grouped:
        invoices_grouped[key] = {
            'date': t['date'],
            'customer_code': t['customer_code'],
            'items': []
        }
    invoices_grouped[key]['items'].append(t)

invoices_list = []
nota_counter = 1
for key, group in invoices_grouped.items():
    nota_no = f"INV-{group['date'].replace('-', '')}-{nota_counter:03d}"
    nota_counter += 1
    total_amount = sum(item['subtotal'] for item in group['items'])
    invoices_list.append({
        'invoice_no': nota_no,
        'date': group['date'],
        'customer_code': group['customer_code'],
        'total_amount': total_amount,
        'items': group['items']
    })

seed_data = {
    'customers': customers_list,
    'products': products_list,
    'pricing_matrix': pricing_matrix,
    'stock': stock_list,
    'invoices': invoices_list,
    'transactions_raw': transactions_raw
}

with open('backend/seed_data.json', 'w', encoding='utf-8') as f:
    json.dump(seed_data, f, indent=2, ensure_ascii=False)

print(f"Extraction successful!")
print(f" - Customers: {len(customers_list)}")
print(f" - Products: {len(products_list)}")
print(f" - Pricing Matrix Entries: {len(pricing_matrix)}")
print(f" - Invoices created: {len(invoices_list)}")
print(f" - Raw Transaction rows: {len(transactions_raw)}")
print(f"Saved to backend/seed_data.json")
