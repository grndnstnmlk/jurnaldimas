import openpyxl
import json

wb = openpyxl.load_workbook('JURNAL KEUANGAN (fix).xlsm', data_only=True)
wb_formula = openpyxl.load_workbook('JURNAL KEUANGAN (fix).xlsm', data_only=False)

info = {}

# 1. Master Data Sheet
ws_md = wb['Master Data']
customers = []
for c in range(3, ws_md.max_column + 1):
    val = ws_md.cell(2, c).value
    if val and str(val).strip() not in ['Total Produk', '120']:
        customers.append(str(val).strip())

products = []
product_prices = {} # prod -> {cust_code: price}
for r in range(3, ws_md.max_row + 1):
    prod = ws_md.cell(r, 2).value
    if prod:
        prod_name = str(prod).strip()
        products.append(prod_name)
        product_prices[prod_name] = {}
        for idx, cust in enumerate(customers):
            c_col = idx + 3
            val = ws_md.cell(r, c_col).value
            product_prices[prod_name][cust] = val

# 2. Data Produk Sheet (Cost / Modal)
ws_dp = wb['Data Produk']
product_cost = {} # prod -> cost
for r in range(9, ws_dp.max_row + 1):
    prod = ws_dp.cell(r, 1).value
    cost = ws_dp.cell(r, 2).value
    if prod:
        prod_name = str(prod).strip()
        product_cost[prod_name] = cost

# 3. Customer Mapping from Nota Transaksi
ws_nota = wb['Nota Transaksi']
customer_map = {} # code -> full name
for r in range(9, 50):
    full_name = ws_nota.cell(r, 10).value
    code = ws_nota.cell(r, 11).value
    if full_name or code:
        customer_map[str(code).strip() if code else ''] = str(full_name).strip() if full_name else ''

# 4. Transactions in Nota Transaksi
transactions = []
for r in range(11, ws_nota.max_row + 1):
    tgl = ws_nota.cell(r, 2).value
    cust = ws_nota.cell(r, 3).value
    prod = ws_nota.cell(r, 4).value
    qty = ws_nota.cell(r, 5).value
    subtotal = ws_nota.cell(r, 6).value
    if any([tgl, cust, prod, qty, subtotal]):
        transactions.append({
            'row': r,
            'date': str(tgl),
            'customer': str(cust).strip() if cust else '',
            'product': str(prod).strip() if prod else '',
            'qty': qty,
            'subtotal': subtotal
        })

# 5. Laba Rugi
ws_lr = wb['Laba-Rugi']
lr_data = []
for r in range(5, ws_lr.max_row + 1):
    prod = ws_lr.cell(r, 2).value
    qty = ws_lr.cell(r, 3).value
    modal = ws_lr.cell(r, 4).value
    jual = ws_lr.cell(r, 5).value
    laba = ws_lr.cell(r, 6).value
    if prod:
        lr_data.append({
            'product': str(prod).strip(),
            'qty': qty,
            'modal': modal,
            'jual': jual,
            'laba': laba
        })

# 6. Stock Opname
ws_so = wb['Stock Opname']
stock_data = []
for r in range(7, ws_so.max_row + 1):
    prod = ws_so.cell(r, 2).value
    awal = ws_so.cell(r, 3).value
    in_qty = ws_so.cell(r, 4).value
    out_qty = ws_so.cell(r, 5).value
    akhir = ws_so.cell(r, 6).value
    if prod:
        stock_data.append({
            'product': str(prod).strip(),
            'awal': awal,
            'in': in_qty,
            'out': out_qty,
            'akhir': akhir
        })

# Check Stock In / Out log in Stock Opname sheet
stock_in_logs = []
for r in range(7, ws_so.max_row + 1):
    tgl = ws_so.cell(r, 8).value
    prod = ws_so.cell(r, 9).value
    qty = ws_so.cell(r, 10).value
    ket = ws_so.cell(r, 11).value
    if any([tgl, prod, qty, ket]):
        stock_in_logs.append({
            'date': str(tgl),
            'product': str(prod).strip() if prod else '',
            'qty': qty,
            'ket': str(ket).strip() if ket else ''
        })

print(f"Total customers in MD: {len(customers)}")
print(f"Customer map: {customer_map}")
print(f"Total products: {len(products)}")
print(f"Total transactions: {len(transactions)}")
print(f"Total Laba Rugi rows: {len(lr_data)}")
print(f"Total Stock items: {len(stock_data)}")
print(f"Total Stock In logs: {len(stock_in_logs)}")

# Check formulas in Nota Transaksi Subtotal and Laba Rugi
print("\n--- Formulas Check ---")
print("Nota Row 11 Subtotal Formula:", wb_formula['Nota Transaksi'].cell(11, 6).value)
print("Nota Row 8 Total Formula:", wb_formula['Nota Transaksi'].cell(8, 5).value)
print("Laba Rugi Row 5 Modal Formula:", wb_formula['Laba-Rugi'].cell(5, 4).value)
print("Laba Rugi Row 5 Jual Formula:", wb_formula['Laba-Rugi'].cell(5, 5).value)
print("Laba Rugi Row 5 Laba Formula:", wb_formula['Laba-Rugi'].cell(5, 6).value)
print("Stock Opname Row 7 In Formula:", wb_formula['Stock Opname'].cell(7, 4).value)
print("Stock Opname Row 7 Out Formula:", wb_formula['Stock Opname'].cell(7, 5).value)
print("Stock Opname Row 7 Akhir Formula:", wb_formula['Stock Opname'].cell(7, 6).value)
